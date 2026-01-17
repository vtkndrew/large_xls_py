"""
Оптимальная вставка строк в Excel файлы с сохранением форматирования и обновлением ссылок.

Гибридный подход: pandas для быстрой вставки данных + openpyxl для форматирования.
Поддерживает межлистовые ссылки и гиперссылки.
"""

from openpyxl import load_workbook
from openpyxl.styles import Font, Fill, Border, Alignment, Protection
from openpyxl.worksheet.hyperlink import Hyperlink
import pandas as pd
from copy import copy
import os
import re
import shutil
from typing import Dict, List, Tuple, Any


# ===== УТИЛИТАРНЫЕ ФУНКЦИИ =====

def validate_input(excel_file: str, target_sheet_name: str, insert_data: List[Dict],
                   dt_column_index: int, kt_column_index: int) -> None:
    """
    Проверяет корректность входных данных.
    
    Args:
        excel_file: путь к Excel файлу
        target_sheet_name: имя целевого листа
        insert_data: массив данных для вставки
        dt_column_index: индекс столбца DT (0-based)
        kt_column_index: индекс столбца KT (0-based)
        
    Raises:
        FileNotFoundError: если файл не найден
        ValueError: если данные некорректны
    """
    # Проверка индексов столбцов
    if dt_column_index < 0:
        raise ValueError(f"dt_column_index должен быть >= 0, получено: {dt_column_index}")
    if kt_column_index < 0:
        raise ValueError(f"kt_column_index должен быть >= 0, получено: {kt_column_index}")
    
    # Проверка существования файла
    if not os.path.exists(excel_file):
        raise FileNotFoundError(f"Файл не найден: {excel_file}")
    
    # Проверка формата данных
    if not isinstance(insert_data, list):
        raise ValueError("insert_data должен быть списком")
    
    if len(insert_data) == 0:
        raise ValueError("insert_data не может быть пустым")
    
    # Проверка каждого элемента
    for idx, item in enumerate(insert_data):
        if not isinstance(item, dict):
            raise ValueError(f"Элемент {idx} в insert_data должен быть словарём")
        
        if 'row_number' not in item:
            raise ValueError(f"Элемент {idx}: отсутствует поле 'row_number'")
        
        if 'new_rows' not in item:
            raise ValueError(f"Элемент {idx}: отсутствует поле 'new_rows'")
        
        if not isinstance(item['new_rows'], list):
            raise ValueError(f"Элемент {idx}: 'new_rows' должен быть списком")
        
        if len(item['new_rows']) == 0:
            raise ValueError(f"Элемент {idx}: 'new_rows' не может быть пустым")
        
        for row_idx, new_row in enumerate(item['new_rows']):
            if 'dt' not in new_row:
                raise ValueError(f"Элемент {idx}, строка {row_idx}: отсутствует 'dt'")
            if 'kt' not in new_row:
                raise ValueError(f"Элемент {idx}, строка {row_idx}: отсутствует 'kt'")
    
    # Проверка существования листа
    wb = load_workbook(excel_file, read_only=True, data_only=False)
    if target_sheet_name not in wb.sheetnames:
        wb.close()
        raise ValueError(f"Лист '{target_sheet_name}' не найден в файле. Доступные листы: {wb.sheetnames}")
    wb.close()
    
    print("✓ Валидация пройдена")


def adjust_formula_references(formula: str, source_row: int, target_row: int) -> str:
    """
    Корректирует относительные ссылки в формуле при копировании строки.
    
    Пример:
        formula = "=SUM(A1:A10)"
        source_row = 5
        target_row = 8
        Result: "=SUM(A4:A13)"  (смещение +3)
        
    Args:
        formula: исходная формула
        source_row: номер исходной строки
        target_row: номер целевой строки
        
    Returns:
        Скорректированная формула
    """
    row_diff = target_row - source_row
    
    if row_diff == 0:
        return formula
    
    # Паттерн для поиска ссылок на ячейки
    # Ищем: буквы колонки + цифры строки, но не абсолютные ссылки
    # Негативный lookbehind (?<!\$) - перед колонкой и строкой не должно быть $
    pattern = r'(?<!\$)([A-Z]+)(?<!\$)(\d+)'
    
    def adjust_reference(match):
        col = match.group(1)
        row = int(match.group(2))
        new_row = row + row_diff
        return f"{col}{new_row}"
    
    # Заменяем только относительные ссылки
    adjusted_formula = re.sub(pattern, adjust_reference, formula)
    
    return adjusted_formula


def calculate_cumulative_shifts(row_shift_map: Dict[int, int]) -> Dict[int, int]:
    """
    Создаёт кумулятивную карту смещений строк.
    
    ВАЖНО: Ключи - это ОРИГИНАЛЬНЫЕ номера строк (до вставок).
    Значения - это смещение которое нужно применить к этим строкам.
    
    Пример:
        row_shift_map = {3: 2, 4: 1}
        # После строки 3 вставлено 2, после строки 4 вставлено 1
        
        Оригинал: 1, 2, 3, 4, 5, 6, 7, ...
        После вставок: 1, 2, 3, [4,5 от 3], 4->6, [7 от 4], 5->8, 6->9, 7->10, ...
        
        Result:
        {
            1: 0,  # строка 1 не смещается
            2: 0,  # строка 2 не смещается
            3: 0,  # строка 3 не смещается (сама точка вставки)
            4: 2,  # строка 4 смещается на +2 (становится 6)
            5: 3,  # строка 5 смещается на +3 (становится 8)
            6: 3,  # строка 6 смещается на +3 (становится 9)
            7: 3,  # строка 7 смещается на +3 (становится 10)
            ...
        }
        
    Args:
        row_shift_map: словарь {номер_строки_В_ОРИГИНАЛЕ: количество_вставленных_после}
        
    Returns:
        Кумулятивная карта смещений {оригинальная_строка: смещение}
    """
    cumulative_shifts = {}
    
    if not row_shift_map:
        return cumulative_shifts
    
    # Сортируем точки вставки
    sorted_rows = sorted(row_shift_map.keys())
    
    # Обрабатываем в прямом порядке
    cumulative = 0
    for orig_row in range(1, 100000):
        # Записываем смещение для этой оригинальной строки
        cumulative_shifts[orig_row] = cumulative
        
        # Проверяем, есть ли вставка после этой строки
        if orig_row in row_shift_map:
            # После этой строки вставлено N строк
            cumulative += row_shift_map[orig_row]
        
        # Оптимизация: если прошли все точки вставки
        if orig_row > max(sorted_rows) + 1000:
            # Заполняем остаток константным смещением
            for remaining_orig in range(orig_row + 1, 100000):
                cumulative_shifts[remaining_orig] = cumulative
            break
    
    return cumulative_shifts


def update_formula_with_shifts(formula: str, target_sheet_name: str,
                               cumulative_shifts: Dict[int, int]) -> str:
    """
    Обновляет формулу с учётом смещений строк на целевом листе.
    
    Пример:
        formula = "=Sheet1!A5 + Sheet1!B10"
        cumulative_shifts = {6: 3, 11: 5}
        Result: "=Sheet1!A8 + Sheet1!B15"
        
    Args:
        formula: исходная формула
        target_sheet_name: имя целевого листа
        cumulative_shifts: карта смещений
        
    Returns:
        Обновлённая формула
    """
    # Паттерн для поиска ссылок на целевой лист
    # Поддерживает: Sheet1!A5, 'Sheet Name'!A5, Sheet1!$A$5
    # НЕ экранируем пробелы, т.к. внутри кавычек они обычные
    # Экранируем только спецсимволы regex
    escaped_sheet_name = re.escape(target_sheet_name).replace(r'\ ', ' ')
    pattern = rf"((?:'{escaped_sheet_name}'|{escaped_sheet_name})!)(\$?)([A-Z]+)(\$?)(\d+)"
    
    def replace_reference(match):
        prefix = match.group(1)      # "Sheet1!" или "'Sheet Name'!"
        col_abs = match.group(2)     # "$" или ""
        col = match.group(3)         # "A"
        row_abs = match.group(4)     # "$" или ""
        row_num = int(match.group(5))  # "5"
        
        # Не обновляем абсолютные ссылки на строки
        if row_abs == '$':
            return match.group(0)
        
        # Применяем смещение
        shift = cumulative_shifts.get(row_num, 0)
        new_row_num = row_num + shift
        
        return f"{prefix}{col_abs}{col}{row_abs}{new_row_num}"
    
    new_formula = re.sub(pattern, replace_reference, formula)
    return new_formula


def update_hyperlink_with_shifts(hyperlink_target: str, target_sheet_name: str,
                                 cumulative_shifts: Dict[int, int]) -> str:
    """
    Обновляет гиперссылку с учётом смещений строк.
    
    Пример:
        hyperlink_target = "#Sheet1!A5" или "Sheet1!A5"
        cumulative_shifts = {6: 3}
        Result: "#Sheet1!A8" или "Sheet1!A8"
        
    Args:
        hyperlink_target: целевая ссылка гиперссылки
        target_sheet_name: имя целевого листа
        cumulative_shifts: карта смещений
        
    Returns:
        Обновлённая ссылка
    """
    # НЕ экранируем пробелы, т.к. внутри кавычек они обычные
    escaped_sheet_name = re.escape(target_sheet_name).replace(r'\ ', ' ')
    # Поддержка гиперссылок с # и без
    pattern = rf"(#?(?:'{escaped_sheet_name}'|{escaped_sheet_name})!)(\$?)([A-Z]+)(\$?)(\d+)"
    
    def replace_reference(match):
        prefix = match.group(1)
        col_abs = match.group(2)
        col = match.group(3)
        row_abs = match.group(4)
        row_num = int(match.group(5))
        
        if row_abs == '$':
            return match.group(0)
        
        shift = cumulative_shifts.get(row_num, 0)
        new_row_num = row_num + shift
        
        return f"{prefix}{col_abs}{col}{row_abs}{new_row_num}"
    
    new_target = re.sub(pattern, replace_reference, hyperlink_target)
    return new_target
# ===== ВСПОМОГАТЕЛЬНЫЕ ФУНКЦИИ ДЛЯ V2 =====




def reverse_map_row(target_row: int, cumulative_shifts: Dict[int, int]) -> int:
    """
    Находит номер строки в оригинальном файле по номеру в результирующем.
    
    ВАЖНО: cumulative_shifts теперь использует оригинальные строки как ключи.
    cumulative_shifts[orig_row] = shift, где orig_row + shift = result_row
    
    Пример:
        cumulative_shifts = {1: 0, 2: 0, 3: 0, 4: 2, 5: 3, 6: 3, 7: 3, ...}
        target_row = 6  -> source_row = 4  (4 + 2 = 6)
        target_row = 10 -> source_row = 7  (7 + 3 = 10)
        
    Args:
        target_row: номер строки в результирующем файле
        cumulative_shifts: карта смещений {orig_row: shift}
        
    Returns:
        Номер строки в оригинальном файле
    """
    # Ищем оригинальную строку, для которой orig_row + shift = target_row
    for orig_row, shift in cumulative_shifts.items():
        if orig_row + shift == target_row:
            return orig_row
    
    # Если не нашли (строка без смещения), то target_row = orig_row
    return target_row


def get_inserted_rows_set(row_shift_map: Dict[int, int]) -> set:
    """
    Возвращает множество номеров вставленных строк.
    
    Пример:
        row_shift_map = {5: 3, 10: 2}
        Result: {6, 7, 8, 14, 15}  (с учётом смещения)
        
    Args:
        row_shift_map: карта вставок
        
    Returns:
        Множество номеров вставленных строк
    """
    inserted_rows = set()
    cumulative = 0
    
    for row_num in sorted(row_shift_map.keys()):
        count = row_shift_map[row_num]
        # Строка row_num сдвинута на cumulative
        actual_position = row_num + cumulative
        
        # Вставленные строки идут сразу после actual_position
        for i in range(1, count + 1):
            inserted_rows.add(actual_position + i)
        
        cumulative += count
    
    return inserted_rows


def get_source_row_for_inserted(target_row: int, row_shift_map: Dict[int, int],
                                cumulative_shifts: Dict[int, int]) -> int:
    """
    Определяет исходную строку для вставленной строки.
    
    Args:
        target_row: номер вставленной строки
        row_shift_map: карта вставок
        cumulative_shifts: кумулятивные смещения
        
    Returns:
        Номер исходной строки (откуда копировали)
    """
    # Для каждой точки вставки проверяем, попадает ли target_row в её диапазон
    cumulative = 0
    for row_num in sorted(row_shift_map.keys()):
        count = row_shift_map[row_num]
        actual_position = row_num + cumulative
        
        # Проверяем, находится ли target_row в диапазоне вставленных строк
        if actual_position < target_row <= actual_position + count:
            return row_num
        
        cumulative += count
    
    return 0  # Не должно произойти


def copy_row_formatting(ws_source, source_row: int, ws_target, target_row: int) -> None:
    """
    Копирует форматирование и формулы всех ячеек строки из source в target.
    
    Args:
        ws_source: исходный worksheet
        source_row: номер исходной строки
        ws_target: целевой worksheet
        target_row: номер целевой строки
    """
    # Высота строки
    if ws_source.row_dimensions[source_row].height:
        ws_target.row_dimensions[target_row].height = \
            ws_source.row_dimensions[source_row].height
    
    # Проход по всем ячейкам (используем max_column из source)
    for col_idx in range(1, ws_source.max_column + 1):
        source_cell = ws_source.cell(row=source_row, column=col_idx)
        target_cell = ws_target.cell(row=target_row, column=col_idx)
        
        # Копируем форматирование
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format
        target_cell.protection = copy(source_cell.protection)
        
        # Копируем формулу (pandas их теряет!)
        if source_cell.value and isinstance(source_cell.value, str) \
           and source_cell.value.startswith('='):
            target_cell.value = source_cell.value
        
        # Копируем гиперссылку
        if source_cell.hyperlink:
            target_cell.hyperlink = Hyperlink(target=source_cell.hyperlink.target)


def copy_row_formatting_from_metadata(ws_target, target_row: int,
                                     source_metadata: Dict) -> None:
    """
    Копирует форматирование из сохранённых метаданных.
    
    Args:
        ws_target: целевой worksheet
        target_row: номер целевой строки
        source_metadata: метаданные исходной строки
    """
    # Высота строки
    if source_metadata['row_height']:
        ws_target.row_dimensions[target_row].height = source_metadata['row_height']
    
    # Форматирование ячеек
    for col_idx, formatting in source_metadata['formatting'].items():
        target_cell = ws_target.cell(row=target_row, column=col_idx)
        
        target_cell.font = copy(formatting['font'])
        target_cell.fill = copy(formatting['fill'])
        target_cell.border = copy(formatting['border'])
        target_cell.alignment = copy(formatting['alignment'])
        target_cell.number_format = formatting['number_format']
        target_cell.protection = copy(formatting['protection'])
        
        # Гиперссылки
        if 'hyperlink' in formatting:
            target_cell.hyperlink = Hyperlink(target=formatting['hyperlink'])


def copy_sheet_properties(ws_source, ws_target) -> None:
    """
    Копирует свойства листа: ширину столбцов, фильтры, цвет вкладки и т.д.
    
    Args:
        ws_source: исходный worksheet
        ws_target: целевой worksheet
    """
    # Ширина столбцов
    for col_letter in ws_source.column_dimensions:
        if ws_source.column_dimensions[col_letter].width:
            ws_target.column_dimensions[col_letter].width = \
                ws_source.column_dimensions[col_letter].width
    
    # Автофильтры
    if ws_source.auto_filter and ws_source.auto_filter.ref:
        ws_target.auto_filter.ref = ws_source.auto_filter.ref
    
    # Закреплённые области
    if ws_source.freeze_panes:
        ws_target.freeze_panes = ws_source.freeze_panes
    
    # Цвет вкладки листа
    if ws_source.sheet_properties and ws_source.sheet_properties.tabColor:
        ws_target.sheet_properties.tabColor = ws_source.sheet_properties.tabColor


# ===== ФАЗА 1: ИЗВЛЕЧЕНИЕ МЕТАДАННЫХ =====

def extract_target_sheet_metadata(wb, target_sheet_name: str, 
                                  row_numbers: List[int]) -> Dict[int, Dict]:
    """
    Извлекает форматирование и формулы исходных строк с целевого листа.
    
    Args:
        wb: загруженный workbook
        target_sheet_name: имя целевого листа
        row_numbers: список номеров строк для копирования
        
    Returns:
        Словарь метаданных для каждой строки
    """
    ws = wb[target_sheet_name]
    metadata = {}
    
    for row_num in row_numbers:
        row_data = {
            'formatting': {},
            'formulas': {},
            'row_height': ws.row_dimensions[row_num].height,
            'merged_cells': []
        }
        
        # Проверяем merged cells в этой строке
        for merged_range in ws.merged_cells.ranges:
            if row_num >= merged_range.min_row and row_num <= merged_range.max_row:
                row_data['merged_cells'].append(str(merged_range))
        
        # Проход по всем ячейкам строки
        for cell in ws[row_num]:
            col_idx = cell.column
            
            # Сохраняем форматирование
            row_data['formatting'][col_idx] = {
                'font': copy(cell.font),
                'fill': copy(cell.fill),
                'border': copy(cell.border),
                'alignment': copy(cell.alignment),
                'number_format': cell.number_format,
                'protection': copy(cell.protection)
            }
            
            # Сохраняем формулы
            if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                row_data['formulas'][col_idx] = cell.value
            
            # Сохраняем гиперссылки
            if cell.hyperlink:
                row_data['formatting'][col_idx]['hyperlink'] = cell.hyperlink.target
        
        metadata[row_num] = row_data
    
    return metadata


# ===== ФАЗА 1.5: АНАЛИЗ МЕЖЛИСТОВЫХ ССЫЛОК =====

def analyze_cross_sheet_references(wb, target_sheet_name: str,
                                   insert_positions: List[Tuple[int, int]]) -> Dict:
    """
    Находит все формулы и гиперссылки на других листах, ссылающиеся на целевой лист.
    
    Args:
        wb: загруженный workbook
        target_sheet_name: имя целевого листа
        insert_positions: список кортежей (row_number, count)
        
    Returns:
        Словарь ссылок для обновления
    """
    cross_references = {}
    
    # Паттерн для поиска ссылок на целевой лист
    # НЕ экранируем пробелы, т.к. внутри кавычек они обычные
    escaped_sheet_name = re.escape(target_sheet_name).replace(r'\ ', ' ')
    pattern = rf"(?:'{escaped_sheet_name}'|{escaped_sheet_name})!(\$?)([A-Z]+)(\$?)(\d+)"
    
    # Проход по всем листам кроме целевого
    for sheet_name in wb.sheetnames:
        if sheet_name == target_sheet_name:
            continue
        
        ws = wb[sheet_name]
        
        # Проход по всем ячейкам листа
        for row in ws.iter_rows():
            for cell in row:
                cell_references = []
                
                # Проверка формул
                if cell.value and isinstance(cell.value, str) and cell.value.startswith('='):
                    formula = cell.value
                    matches = list(re.finditer(pattern, formula))
                    
                    if matches:
                        cell_references.append({
                            'type': 'formula',
                            'original_formula': formula
                        })
                
                # Проверка гиперссылок
                if cell.hyperlink:
                    target = cell.hyperlink.target
                    if target:
                        # Проверяем внутренние ссылки (с # и без)
                        # Ищем ссылки на целевой лист
                        match = re.search(pattern, target)
                        if match:
                            cell_references.append({
                                'type': 'hyperlink',
                                'original_target': target
                            })
                
                # Сохраняем ссылки для обновления
                if cell_references:
                    key = f"{sheet_name}!{cell.coordinate}"
                    cross_references[key] = {
                        'sheet': sheet_name,
                        'row': cell.row,
                        'col': cell.column,
                        'references': cell_references
                    }
    
    return cross_references


# ===== ФАЗА 2: ВСТАВКА ЧЕРЕЗ PANDAS =====

def insert_rows_with_pandas(excel_file: str, target_sheet_name: str,
                            insert_data: List[Dict],
                            dt_column_index: int, kt_column_index: int) -> Tuple[str, List[Dict], Dict[int, int]]:
    """
    Быстрая вставка строк через pandas.
    
    Args:
        excel_file: путь к Excel файлу
        target_sheet_name: имя целевого листа
        insert_data: данные для вставки
        dt_column_index: индекс столбца DT (0-based)
        kt_column_index: индекс столбца KT (0-based)
        
    Returns:
        Кортеж: (путь к временному файлу, список вставленных позиций, карта смещений)
    """
    # Загрузка только целевого листа БЕЗ заголовков
    # header=None предотвращает появление "Unnamed: X" для пустых ячеек в первой строке
    df = pd.read_excel(excel_file, sheet_name=target_sheet_name, engine='openpyxl', header=None)
    
    # Сортировка в обратном порядке (от конца к началу)
    sorted_inserts = sorted(insert_data, key=lambda x: x['row_number'], reverse=True)
    
    row_shift_map = {}
    
    for item in sorted_inserts:
        row_idx = item['row_number']
        pandas_row_idx = row_idx - 1  # pandas использует 0-based индексацию
        new_rows_data = item['new_rows']
        
        # Получаем исходную строку как шаблон
        source_row = df.iloc[pandas_row_idx].copy()
        
        # Создаём новые строки
        new_rows_list = []
        for new_row_info in new_rows_data:
            new_row = source_row.copy()
            
            # Обновляем значения в столбцах DT и KT используя индексы
            # С header=None столбцы имеют числовые индексы (0, 1, 2, ...)
            if len(df.columns) > dt_column_index:
                new_row[dt_column_index] = new_row_info['dt']
            if len(df.columns) > kt_column_index:
                new_row[kt_column_index] = new_row_info['kt']
            
            new_rows_list.append(new_row)
        
        # Вставка строк после исходной строки
        df_new_rows = pd.DataFrame(new_rows_list)
        df = pd.concat([
            df.iloc[:pandas_row_idx + 1],
            df_new_rows,
            df.iloc[pandas_row_idx + 1:]
        ], ignore_index=True)
        
        # Обновляем карту смещений
        num_inserted = len(new_rows_data)
        row_shift_map[row_idx] = num_inserted
    
    # ВАЖНО: Пересчитываем inserted_positions ПОСЛЕ всех вставок
    # Обрабатываем в прямом порядке с учётом кумулятивных смещений
    inserted_positions = []
    cumulative = 0
    for row_num in sorted(row_shift_map.keys()):
        count = row_shift_map[row_num]
        actual_position = row_num + cumulative
        
        # Вставленные строки идут сразу после actual_position
        for i in range(count):
            inserted_positions.append({
                'row_number': actual_position + 1 + i,  # Excel 1-based
                'source_row': row_num
            })
        
        cumulative += count
    
    # Создание временного файла
    temp_file = excel_file.replace('.xlsx', '_temp.xlsx')
    
    # Копируем весь исходный файл
    shutil.copy(excel_file, temp_file)
    
    # Перезаписываем только целевой лист БЕЗ заголовков
    with pd.ExcelWriter(temp_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        df.to_excel(writer, sheet_name=target_sheet_name, index=False, header=False)
    
    return temp_file, inserted_positions, row_shift_map


# ===== ФАЗА 3: ПРИМЕНЕНИЕ ФОРМАТИРОВАНИЯ V2 =====

def apply_formatting_to_target_sheet_v2(
    original_file: str,
    temp_file: str,
    target_sheet_name: str,
    row_shift_map: Dict[int, int],
    metadata: Dict[int, Dict],
    inserted_positions: List[Dict],
    dt_column_index: int,
    kt_column_index: int
) -> None:
    """
    Восстанавливает ВСЁ форматирование и формулы из оригинального файла.
    
    V2: Копирует форматирование ВСЕХ строк из оригинала, а не только вставленных.
    Это решает проблему потери форматирования при сохранении через pandas.
    
    Args:
        original_file: путь к оригинальному файлу
        temp_file: путь к временному файлу (после pandas)
        target_sheet_name: имя целевого листа
        row_shift_map: карта смещений строк
        metadata: метаданные для вставленных строк
        inserted_positions: список позиций вставленных строк
        dt_column_index: индекс столбца DT (0-based, будет преобразован в 1-based)
        kt_column_index: индекс столбца KT (0-based, будет преобразован в 1-based)
    """
    print("  Открываем оригинальный файл для копирования форматирования...")
    
    # Открываем оригинал в обычном режиме (эффективнее чем сохранять всё в метаданные)
    wb_original = load_workbook(original_file, data_only=False)
    wb_temp = load_workbook(temp_file, data_only=False)
    
    ws_original = wb_original[target_sheet_name]
    ws_temp = wb_temp[target_sheet_name]
    
    # Создаём карты
    cumulative_shifts = calculate_cumulative_shifts(row_shift_map)
    inserted_rows = get_inserted_rows_set(row_shift_map)
    
    max_row_original = ws_original.max_row
    max_row_temp = ws_temp.max_row
    
    print(f"  Копирование форматирования для {max_row_temp} строк...")
    
    # Проход по всем строкам результирующего файла
    for target_row in range(1, max_row_temp + 1):
        
        if target_row in inserted_rows:
            # Это новая вставленная строка - используем метаданные
            source_row = get_source_row_for_inserted(target_row, row_shift_map, cumulative_shifts)
            
            if source_row in metadata:
                copy_row_formatting_from_metadata(ws_temp, target_row, metadata[source_row])
                
                # Восстанавливаем формулы для вставленных строк (кроме DT и KT)
                # Копируем БЕЗ корректировки - формулы остаются как в оригинале
                # Преобразуем 0-based индексы в 1-based для openpyxl
                dt_col_openpyxl = dt_column_index + 1
                kt_col_openpyxl = kt_column_index + 1
                
                for col_idx, formula in metadata[source_row]['formulas'].items():
                    if col_idx not in [dt_col_openpyxl, kt_col_openpyxl]:  # Не трогаем столбцы DT и KT
                        target_cell = ws_temp.cell(row=target_row, column=col_idx)
                        target_cell.value = formula  # Копируем как есть, без изменений
        else:
            # Оригинальная строка (возможно сдвинутая)
            source_row = reverse_map_row(target_row, cumulative_shifts)
            
            if source_row > 0 and source_row <= max_row_original:
                # Копируем форматирование и формулы из оригинала
                copy_row_formatting(ws_original, source_row, ws_temp, target_row)
    
    # Копируем свойства листа (ширина столбцов, фильтры и т.д.)
    print("  Копирование свойств листа...")
    copy_sheet_properties(ws_original, ws_temp)
    
    # Закрываем оригинал
    wb_original.close()
    
    # Сохраняем изменения во временном файле
    print("  Сохранение изменений...")
    wb_temp.save(temp_file)
    wb_temp.close()


# ===== ФАЗА 4: ОБНОВЛЕНИЕ МЕЖЛИСТОВЫХ ССЫЛОК =====

def update_cross_sheet_references(wb, target_sheet_name: str, 
                                  cross_references: Dict, 
                                  row_shift_map: Dict[int, int]) -> int:
    """
    Обновляет все межлистовые ссылки на других листах.
    
    Args:
        wb: загруженный workbook
        target_sheet_name: имя целевого листа
        cross_references: словарь найденных ссылок
        row_shift_map: карта смещений строк
        
    Returns:
        Количество обновлённых ссылок
    """
    # Создаём кумулятивную карту смещений
    cumulative_shifts = calculate_cumulative_shifts(row_shift_map)
    
    updated_count = 0
    
    for ref_key, ref_info in cross_references.items():
        sheet_name = ref_info['sheet']
        cell_row = ref_info['row']
        cell_col = ref_info['col']
        
        ws = wb[sheet_name]
        cell = ws.cell(row=cell_row, column=cell_col)
        
        for reference in ref_info['references']:
            if reference['type'] == 'formula':
                # Обновляем формулу
                old_formula = reference['original_formula']
                new_formula = update_formula_with_shifts(
                    old_formula,
                    target_sheet_name,
                    cumulative_shifts
                )
                
                if new_formula != old_formula:
                    cell.value = new_formula
                    updated_count += 1
            
            elif reference['type'] == 'hyperlink':
                # Обновляем гиперссылку
                old_target = reference['original_target']
                new_target = update_hyperlink_with_shifts(
                    old_target,
                    target_sheet_name,
                    cumulative_shifts
                )
                
                if new_target != old_target:
                    cell.hyperlink = Hyperlink(target=new_target)
                    updated_count += 1
    
    return updated_count


# ===== ГЛАВНАЯ ФУНКЦИЯ =====

def insert_rows_optimized(excel_file: str, target_sheet_name: str,
                         insert_data: List[Dict], output_file: str,
                         dt_column_index: int = 8, kt_column_index: int = 9) -> None:
    """
    Главная функция для оптимальной вставки строк с обновлением ссылок.
    
    Args:
        excel_file: путь к входному Excel файлу
        target_sheet_name: имя листа для вставки строк
        insert_data: массив с данными для вставки в формате:
            [{
                "row_number": 5,
                "new_rows": [
                    {"dt": 123, "kt": 567},
                    ...
                ]
            }, ...]
        output_file: путь к выходному файлу
        dt_column_index: индекс столбца DT (0-based, по умолчанию 8 = колонка I)
        kt_column_index: индекс столбца KT (0-based, по умолчанию 9 = колонка J)
        
    Example:
        insert_rows_optimized(
            "Book1.xlsx",
            "Sheet1",
            [{"row_number": 5, "new_rows": [{"dt": 1, "kt": 2}]}],
            "Book1_expanded.xlsx",
            dt_column_index=8,
            kt_column_index=9
        )
    """
    print("="*60)
    print(f"Обработка файла: {excel_file}")
    print(f"Целевой лист: {target_sheet_name}")
    print("="*60)
    
    # 1. Валидация входных данных
    print("\n[1/8] Валидация входных данных...")
    validate_input(excel_file, target_sheet_name, insert_data, dt_column_index, kt_column_index)
    
    # 2. Загрузка workbook
    print("\n[2/8] Загрузка Excel файла...")
    wb = load_workbook(excel_file, data_only=False)
    print(f"  Найдено листов: {len(wb.sheetnames)}")
    
    # 3. Извлечение номеров строк для копирования
    row_numbers = [item['row_number'] for item in insert_data]
    print(f"  Точек вставки: {len(row_numbers)}")
    
    # 4. Фаза 1: Извлечение метаданных целевого листа
    print("\n[3/8] Извлечение метаданных целевого листа...")
    metadata = extract_target_sheet_metadata(wb, target_sheet_name, row_numbers)
    print(f"  Сохранено метаданных для {len(metadata)} строк")
    
    # 5. Фаза 1.5: Анализ межлистовых ссылок
    print("\n[4/8] Анализ межлистовых ссылок...")
    insert_positions = [(item['row_number'], len(item['new_rows'])) 
                       for item in insert_data]
    cross_references = analyze_cross_sheet_references(wb, target_sheet_name, insert_positions)
    print(f"  Найдено ячеек с ссылками на целевой лист: {len(cross_references)}")
    
    # Закрываем workbook для освобождения памяти
    wb.close()
    
    # 6. Фаза 2: Вставка строк через pandas
    print("\n[5/8] Вставка строк через pandas...")
    temp_file, inserted_positions, row_shift_map = insert_rows_with_pandas(
        excel_file, target_sheet_name, insert_data, dt_column_index, kt_column_index
    )
    total_inserted = sum(len(item['new_rows']) for item in insert_data)
    print(f"  Вставлено строк: {total_inserted}")
    
    # 7. Фаза 3: Восстановление форматирования из оригинала (V2)
    print("\n[6/8] Восстановление форматирования из оригинала...")
    apply_formatting_to_target_sheet_v2(
        excel_file,
        temp_file,
        target_sheet_name,
        row_shift_map,
        metadata,
        inserted_positions,
        dt_column_index,
        kt_column_index
    )
    print(f"  ✓ Форматирование восстановлено для всех строк")
    
    # 8. Фаза 4: Обновление межлистовых ссылок
    print("\n[7/8] Обновление межлистовых ссылок...")
    wb = load_workbook(temp_file, data_only=False)
    updated_count = update_cross_sheet_references(wb, target_sheet_name, cross_references, row_shift_map)
    print(f"  Обновлено ссылок: {updated_count}")
    
    # 9. Сохранение результата
    print(f"\n[8/8] Сохранение результата в {output_file}...")
    wb.save(output_file)
    wb.close()
    
    # 10. Очистка временного файла
    if os.path.exists(temp_file):
        os.remove(temp_file)
    
    print("\n" + "="*60)
    print("✓ Готово!")
    print(f"  Вставлено строк: {total_inserted}")
    print(f"  Обновлено межлистовых ссылок: {updated_count}")
    print(f"  Результат сохранён в: {output_file}")
    print("="*60)


# ===== ПРИМЕР ИСПОЛЬЗОВАНИЯ =====

if __name__ == "__main__":
    # Пример данных для вставки
    insert_data = [
        {
            "row_number": 3,
            "new_rows": [
                {"dt": 2, "kt": 6},
                {"dt": 2, "kt": 5},
                {"dt": 4, "kt": 6},
                {"dt": 4, "kt": 5},
                {"dt": 7, "kt": 6},
                {"dt": 7, "kt": 5},
            ]
        },
        {
            "row_number": 5,
            "new_rows": [
                {"dt": 8, "kt": 3},
                {"dt": 8, "kt": 2},
            ]
        },
        {
            "row_number": 7,
            "new_rows": [
                {"dt": 9, "kt": 4},
                {"dt": 2, "kt": 4},
            ]
        },
        {
            "row_number": 9,
            "new_rows": [
                {"dt": 13, "kt": 12},
                {"dt": 13, "kt": 19},
            ]
        },
        {
            "row_number": 11,
            "new_rows": [
                {"dt": 4, "kt": 8},
                {"dt": 4, "kt": 5},
                {"dt": 4, "kt": 12},
            ]
        },
    ]
    
    try:
        insert_rows_optimized(
            excel_file="Book1.xlsx",
            target_sheet_name="Some Custom Sheet1",
            insert_data=insert_data,
            output_file="Book1_expanded_NEW.xlsx",
            dt_column_index=1,  # Колонка I (0-based)
            kt_column_index=2   # Колонка J (0-based)
        )
    except Exception as e:
        print(f"\n❌ Ошибка: {e}")
        import traceback
        traceback.print_exc()
